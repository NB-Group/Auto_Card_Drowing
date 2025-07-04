#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
春秋杀 - 游戏记录表生成器
生成游戏过程中使用的空白记录表，供玩家手写记录每轮变化
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import glob
import os

class GameRecordSheet:
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "春秋杀游戏记录表"
        
        # 样式定义
        self.header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        self.label_font = Font(name='微软雅黑', size=10, bold=True)
        self.data_font = Font(name='微软雅黑', size=9)
        self.small_font = Font(name='微软雅黑', size=8)
        
        # 填充色
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.label_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        self.round_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # 边框样式
        self.thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 对齐样式
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
    def setup_game_info(self):
        """设置游戏基本信息区域"""
        # 主标题
        self.worksheet.merge_cells('A1:M1')
        title_cell = self.worksheet['A1']
        title_cell.value = "🏰 春秋杀游戏记录表"
        title_cell.font = Font(name='微软雅黑', size=18, bold=True, color='2F5597')
        title_cell.alignment = self.center_alignment
        title_cell.fill = self.header_fill
        title_cell.border = self.thick_border
        
        # 游戏信息
        row = 3
        info_data = [
            ("📅 游戏日期:", ""),
            ("🎮 玩家数量:", ""),
            ("⏰ 开始时间:", ""),
            ("🎯 游戏轮数:", ""),
        ]
        
        for i, (label, value) in enumerate(info_data):
            # 标签
            label_cell = self.worksheet.cell(row=row, column=1 + i*3)
            label_cell.value = label
            label_cell.font = self.label_font
            label_cell.fill = self.label_fill
            label_cell.border = self.thin_border
            
            # 空白填写区域（合并2列）
            self.worksheet.merge_cells(f'{get_column_letter(2 + i*3)}{row}:{get_column_letter(3 + i*3)}{row}')
            value_cell = self.worksheet.cell(row=row, column=2 + i*3)
            value_cell.border = self.thin_border
    
    def setup_player_info(self):
        """设置玩家初始信息区域"""
        start_row = 5
        
        # 标题
        self.worksheet.merge_cells(f'A{start_row}:M{start_row}')
        title_cell = self.worksheet[f'A{start_row}']
        title_cell.value = "👥 玩家初始信息（8人）"
        title_cell.font = self.header_font
        title_cell.fill = self.header_fill
        title_cell.alignment = self.center_alignment
        title_cell.border = self.thick_border
        
        # 表头
        headers = [
            "🎭 玩家姓名",
            "🏛️ 选择国家", 
            "⚔️ 军事力",
            "💰 经济力",
            "🎯 政治影响",
            "❤️ 生命力",
            "🪙 春秋币",
            "🤝 结盟状态",
            "🌟 特殊状态",
            "📝 备注"
        ]
        
        header_row = start_row + 1
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=header_row, column=col)
            cell.value = header
            cell.font = self.label_font
            cell.fill = self.label_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 8个国家的初始数据参考
        countries_data = [
            "🐉 秦国 (军85/经70/政60)",
            "🐦 楚国 (军75/经80/政70)", 
            "🦅 齐国 (军70/经90/政85)",
            "🕊️ 燕国 (军60/经50/政55)",
            "🐎 赵国 (军80/经65/政70)",
            "🛡️ 魏国 (军75/经75/政80)",
            "🏹 韩国 (军65/经60/政65)",
            "👑 周王室 (军40/经85/政95)"
        ]
        
        # 预留8个玩家位置
        for player_num in range(1, 9):
            row = header_row + player_num
            for col in range(1, len(headers) + 1):
                cell = self.worksheet.cell(row=row, column=col)
                if col == 1:  # 玩家姓名列
                    cell.value = f"玩家{player_num}:"
                    cell.font = self.small_font
                elif col == 2:  # 国家列，显示参考信息
                    if player_num <= len(countries_data):
                        cell.value = countries_data[player_num - 1]
                        cell.font = self.small_font
                cell.border = self.thin_border
        
        return start_row + 10  # 返回下一个区域的起始行
    
    def setup_zhou_tax_record(self, start_row):
        """设置周王室税收记录区域"""
        # 标题
        self.worksheet.merge_cells(f'A{start_row}:M{start_row}')
        title_cell = self.worksheet[f'A{start_row}']
        title_cell.value = "👑 周王室税收记录"
        title_cell.font = self.header_font
        title_cell.fill = self.header_fill
        title_cell.alignment = self.center_alignment
        title_cell.border = self.thick_border
        
        # 税收说明
        tax_info_row = start_row + 1
        self.worksheet.merge_cells(f'A{tax_info_row}:M{tax_info_row}')
        tax_info_cell = self.worksheet[f'A{tax_info_row}']
        tax_info_cell.value = "📋 各国向周王室纳税：秦10、楚8、齐12、燕5、赵8、魏10、韩6春秋币/回合（周王室自己不用纳税）"
        tax_info_cell.font = self.small_font
        tax_info_cell.alignment = self.left_alignment
        tax_info_cell.border = self.thin_border
        
        # 表头
        headers = [
            "🎯 回合数",
            "🐉 秦国",
            "🐦 楚国", 
            "🦅 齐国",
            "🕊️ 燕国",
            "🐎 赵国",
            "🛡️ 魏国",
            "🏹 韩国",
            "👑 周收入",
            "💰 总收入",
            "📝 备注"
        ]
        
        header_row = start_row + 2
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=header_row, column=col)
            cell.value = header
            cell.font = self.label_font
            cell.fill = self.label_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 预留15轮税收记录
        for round_num in range(1, 16):
            row = header_row + round_num
            for col in range(1, len(headers) + 1):
                cell = self.worksheet.cell(row=row, column=col)
                if col == 1:  # 回合数列
                    cell.value = f"第{round_num}轮"
                    cell.font = self.small_font
                    cell.fill = self.round_fill
                cell.border = self.thin_border
        
        return start_row + 18  # 返回下一个区域的起始行
    
    def setup_round_records(self, start_row):
        """设置回合记录区域（每轮每个玩家的状态）"""
        # 标题
        self.worksheet.merge_cells(f'A{start_row}:M{start_row}')
        title_cell = self.worksheet[f'A{start_row}']
        title_cell.value = "🔄 每轮玩家状态记录（每轮结束时填写所有玩家当前状态）"
        title_cell.font = self.header_font
        title_cell.fill = self.header_fill
        title_cell.alignment = self.center_alignment
        title_cell.border = self.thick_border
        
        # 表头
        headers = [
            "🎯 回合/玩家",
            "🏛️ 国家名称",
            "⚔️ 军事力",
            "💰 经济力",
            "🎯 政治影响", 
            "❤️ 生命力",
            "🪙 春秋币",
            "🤝 结盟状态",
            "🌟 特殊加成",
            "📝 备注"
        ]
        
        header_row = start_row + 1
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=header_row, column=col)
            cell.value = header
            cell.font = self.label_font
            cell.fill = self.label_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 预留10轮记录空间，每轮8个玩家
        for round_num in range(1, 11):
            # 每轮开始先标记回合数
            round_start_row = header_row + (round_num - 1) * 9 + 1
            self.worksheet.merge_cells(f'A{round_start_row}:J{round_start_row}')
            round_cell = self.worksheet[f'A{round_start_row}']
            round_cell.value = f"--- 第 {round_num} 轮 ---"
            round_cell.font = self.label_font
            round_cell.fill = self.round_fill
            round_cell.alignment = self.center_alignment
            round_cell.border = self.thick_border
            
            # 8个玩家记录行
            for player_num in range(1, 9):
                row = round_start_row + player_num
                for col in range(1, len(headers) + 1):
                    cell = self.worksheet.cell(row=row, column=col)
                    if col == 1:  # 玩家标识列
                        cell.value = f"玩家{player_num}"
                        cell.font = self.small_font
                        cell.fill = self.label_fill
                    cell.border = self.thin_border
        
        return start_row + 91  # 返回下一个区域的起始行 (1标题 + 1表头 + 10轮*(1标题+8行))
    
    def setup_card_records(self, start_row):
        """设置卡牌记录区域"""
        # 标题
        self.worksheet.merge_cells(f'A{start_row}:M{start_row}')
        title_cell = self.worksheet[f'A{start_row}']
        title_cell.value = "🃏 卡牌购买/使用记录"
        title_cell.font = self.header_font
        title_cell.fill = self.header_fill
        title_cell.alignment = self.center_alignment
        title_cell.border = self.thick_border
        
        # 左侧：卡牌购买记录
        buy_start_col = 1
        self.worksheet.merge_cells(f'{get_column_letter(buy_start_col)}{start_row + 1}:{get_column_letter(buy_start_col + 5)}{start_row + 1}')
        buy_title = self.worksheet.cell(row=start_row + 1, column=buy_start_col)
        buy_title.value = "💰 卡牌购买记录"
        buy_title.font = self.label_font
        buy_title.fill = self.label_fill
        buy_title.alignment = self.center_alignment
        buy_title.border = self.thin_border
        
        buy_headers = ["👤 玩家", "🃏 卡牌名称", "💰 价格", "🎯 回合", "📝 效果", "✅ 状态"]
        for col, header in enumerate(buy_headers):
            cell = self.worksheet.cell(row=start_row + 2, column=buy_start_col + col)
            cell.value = header
            cell.font = self.small_font
            cell.fill = self.label_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 预留15行购买记录
        for i in range(15):
            for col in range(6):
                cell = self.worksheet.cell(row=start_row + 3 + i, column=buy_start_col + col)
                cell.border = self.thin_border
        
        # 右侧：卡牌使用记录
        use_start_col = 7
        self.worksheet.merge_cells(f'{get_column_letter(use_start_col)}{start_row + 1}:{get_column_letter(use_start_col + 5)}{start_row + 1}')
        use_title = self.worksheet.cell(row=start_row + 1, column=use_start_col)
        use_title.value = "🎯 卡牌使用记录"
        use_title.font = self.label_font
        use_title.fill = self.label_fill
        use_title.alignment = self.center_alignment
        use_title.border = self.thin_border
        
        use_headers = ["👤 玩家", "🃏 卡牌名称", "🎯 目标", "🎲 结果", "🔄 回合", "📝 说明"]
        for col, header in enumerate(use_headers):
            cell = self.worksheet.cell(row=start_row + 2, column=use_start_col + col)
            cell.value = header
            cell.font = self.small_font
            cell.fill = self.label_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 预留15行使用记录
        for i in range(15):
            for col in range(6):
                cell = self.worksheet.cell(row=start_row + 3 + i, column=use_start_col + col)
                cell.border = self.thin_border
        
        return start_row + 19  # 返回下一个区域的起始行
    
    def setup_diplomacy_records(self, start_row):
        """设置外交记录区域"""
        # 标题
        self.worksheet.merge_cells(f'A{start_row}:M{start_row}')
        title_cell = self.worksheet[f'A{start_row}']
        title_cell.value = "🤝 外交与结盟记录"
        title_cell.font = self.header_font
        title_cell.fill = self.header_fill
        title_cell.alignment = self.center_alignment
        title_cell.border = self.thick_border
        
        # 表头
        headers = [
            "🎯 回合",
            "👤 发起方",
            "🏛️ 目标方", 
            "🤝 外交行动",
            "💰 条件/代价",
            "✅ 结果",
            "⏰ 有效期",
            "🌟 特殊效果",
            "📝 备注"
        ]
        
        header_row = start_row + 1
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=header_row, column=col)
            cell.value = header
            cell.font = self.label_font
            cell.fill = self.label_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 预留12行外交记录
        for i in range(12):
            for col in range(1, len(headers) + 1):
                cell = self.worksheet.cell(row=header_row + 1 + i, column=col)
                cell.border = self.thin_border
        
        return start_row + 15  # 返回下一个区域的起始行
    
    def setup_notes_area(self, start_row):
        """设置备注区域"""
        # 标题
        self.worksheet.merge_cells(f'A{start_row}:M{start_row}')
        title_cell = self.worksheet[f'A{start_row}']
        title_cell.value = "📝 游戏备注与总结"
        title_cell.font = self.header_font
        title_cell.fill = self.header_fill
        title_cell.alignment = self.center_alignment
        title_cell.border = self.thick_border
        
        # 备注区域（大空白区域）
        notes_area_rows = 8
        for row_offset in range(notes_area_rows):
            self.worksheet.merge_cells(f'A{start_row + 1 + row_offset}:M{start_row + 1 + row_offset}')
            cell = self.worksheet.cell(row=start_row + 1 + row_offset, column=1)
            cell.border = self.thin_border
            if row_offset == 0:
                cell.value = "💡 游戏心得、策略总结、有趣事件等..."
                cell.font = self.small_font
                cell.alignment = self.left_alignment
            
    def adjust_column_widths(self):
        """调整列宽"""
        column_widths = {
            'A': 8,    # 第1列 - 回合/玩家 (很窄)
            'B': 11,   # 第2列 - 国家名称 (紧凑)
            'C': 6,    # 第3列 - 军事力 (数字，很窄)
            'D': 6,    # 第4列 - 经济力 (数字，很窄)
            'E': 6,    # 第5列 - 政治影响 (数字，很窄)
            'F': 6,    # 第6列 - 生命力 (数字，很窄)
            'G': 7,    # 第7列 - 春秋币 (数字，窄)
            'H': 12,   # 第8列 - 结盟状态 (文字，紧凑)
            'I': 12,   # 第9列 - 特殊加成 (文字，紧凑)
            'J': 15,   # 第10列 - 备注 (文字，适中)
            'K': 8,    # 第11列 - 额外数字列 (窄)
            'L': 13,   # 第12列 - 额外文字列 (紧凑)
            'M': 13    # 第13列 - 额外文字列 (紧凑)
        }
        
        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width
    
    def add_game_rules_summary(self, start_row):
        """添加游戏规则简要说明"""
        # 标题
        self.worksheet.merge_cells(f'A{start_row}:M{start_row}')
        title_cell = self.worksheet[f'A{start_row}']
        title_cell.value = "📖 游戏规则速查"
        title_cell.font = self.header_font
        title_cell.fill = self.header_fill
        title_cell.alignment = self.center_alignment
        title_cell.border = self.thick_border
        
        # 规则说明
        rules_text = [
            "💰 经济卡牌(25张)：提升经济力，增加每回合收入",
            "⚔️ 军事卡牌(25张)：提升军事力，增强战斗能力", 
            "👑 周王室税收：其他国家每回合向周王室纳税",
            "🤝 外交结盟：可与其他玩家结盟，获得特殊加成",
            "🎯 胜利条件：军事征服、经济统治或政治影响力最高"
        ]
        
        for i, rule in enumerate(rules_text):
            row = start_row + 1 + i
            self.worksheet.merge_cells(f'A{row}:M{row}')
            cell = self.worksheet.cell(row=row, column=1)
            cell.value = rule
            cell.font = self.small_font
            cell.alignment = self.left_alignment
            cell.border = self.thin_border
            
        return start_row + len(rules_text) + 2
            
    def generate_sheet(self, filename=None):
        """生成游戏记录表"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"春秋杀游戏记录表_{timestamp}.xlsx"
            
        # 依次创建各个区域
        self.setup_game_info()
        
        current_row = self.setup_player_info()
        current_row = self.setup_zhou_tax_record(current_row)
        current_row = self.setup_round_records(current_row)
        current_row = self.setup_card_records(current_row)
        current_row = self.setup_diplomacy_records(current_row)
        current_row = self.add_game_rules_summary(current_row)
        self.setup_notes_area(current_row)
        
        # 调整列宽
        self.adjust_column_widths()
        
        # 保存文件
        self.workbook.save(filename)
        return filename

def main():
    """主函数"""
    print("🎮 春秋杀游戏记录表生成器")
    print("=" * 50)
    
    # 🗑️ 删除旧的游戏记录表文件
    old_files = glob.glob("春秋杀游戏记录表_*.xlsx")
    if old_files:
        print("🗑️ 清理旧文件...")
        for file in old_files:
            try:
                os.remove(file)
                print(f"   ✅ 已删除：{file}")
            except:
                print(f"   ❌ 删除失败：{file}")
        print()
    
    generator = GameRecordSheet()
    filename = generator.generate_sheet()
    
    print(f"✅ 游戏记录表生成成功！")
    print(f"📄 文件路径：{filename}")
    print("📋 包含内容：")
    print("   • 🎮 游戏基本信息填写区")
    print("   • 👥 玩家初始信息区（8个国家参考）")
    print("   • 👑 周王室税收记录区（15轮税收）")
    print("   • 🔄 每轮状态记录区（10轮×8玩家状态）")
    print("   • 🃏 卡牌购买/使用记录区")
    print("   • 🤝 外交结盟记录区")
    print("   • 📖 游戏规则速查")
    print("   • 📝 备注与总结区")
    print("🎯 支持8人游戏，可打印出来手写记录！")

if __name__ == "__main__":
    main() 